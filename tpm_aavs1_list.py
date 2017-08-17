#!/usr/bin/python2.7

import os,sys
from tpm_utils import *
sys.path.append("../board")
sys.path.append("../rf_jig")
sys.path.append("../config")
sys.path.append("../repo_utils")
from bsp.tpm import *
import subprocess
DEVNULL = open(os.devnull,'w')
from ip_scan import *
from optparse import OptionParser
from openpyxl import load_workbook
import matplotlib.pyplot as plt 


RIGHE = 257
COLONNE = 12
COLORS = ['b','g','r','c','m','y','k','w']



def read_aavs1(ws):        
    foglio = []
    for i in range(RIGHE):
        riga = []
        for j in range(COLONNE):
            riga += [ws.cell(row=i+1, column=j+1).value]
        foglio += [riga] 
    #print foglio
    return foglio


def filtro(f, key, val):
    m = []
    nrow = len(f)
    ncol = len(f[0])
    key_cell = 0
    for j in range(ncol):
        if key in f[0][j]:
            key_cell = j
    if key_cell >0: 
        for i in range(nrow):
            if i>0 and val == f[i][key_cell]:
                m += [f[i]] 
    return m

def plot_map(t, marker='o', markersize=12, color='g', print_name=False):
    x = np.transpose(t)[10]
    y = np.transpose(t)[11]
    name = np.transpose(t)[0]
    plt.plot(x,y, marker=marker, markersize=markersize, linestyle = 'None', color=color)
    if print_name:
        for i in range(len(name)):
            plt.annotate("%d"%name[i], xy=(x[i],y[i]), fontsize=10)


if __name__ == "__main__":
    parser = OptionParser()
    parser.add_option("-e", "--excel_file", 
                      dest="ex_file", 
                      default = "",
                      help="The excel file containing the AAVS 1.1 SpreadSheet")
                      
    parser.add_option("-m", "--plot_map", action='store_true',
                      dest="plot_map", 
                      default = False,
                      help="If enabled show the antenna map plot")
                      
    (options, args) = parser.parse_args()

    if not options.ex_file == "" and os.path.isfile(options.ex_file):
        wb2 = load_workbook(options.ex_file, data_only=True)
        ws = wb2.active
        if not ws.title == "AAVS 1.1":
            print "The active Sheet is not what expected (\"%s\" instad of AAVS 1.1)"%(ws.title)
            exit()
    else:
        print "The given argument is not a file!"
        print "Received: %s"%(options.ex_file)
        exit()   

    foglio = read_aavs1(ws)
    wb2.close()

    nrow=len(foglio)
    if nrow>0:
        ncol=len(foglio[0])
    else:
        print "The active sheet is empty."
        exit()

    print "\nExcel File: %s"%(options.ex_file)
    print "\nRow: %d\tColumn: %d"%(nrow, ncol)
    print

    for i in range(nrow):
        print "% 7d\t"%(i),        
        for j in range(ncol):
            #print type(foglio[i][j])
            if type(foglio[i][j])==(unicode or str):
                if len(foglio[i][j])>7:
                    print "%s\t"%(foglio[i][j][:7].center(7,' ')),
                else:
                    print "%s\t"%(foglio[i][j].center(7,' ')),
            elif type(foglio[i][j])==float or type(foglio[i][j])==long:
                if j<9:
                    print "%s\t"%(str("%d"%foglio[i][j]).center(7,' ')),
                else:
                    print "%7.3f\t"%(foglio[i][j]),
            else:
                print " \t",
        print
        if i==0:
            print "--------------------------------------------------------------------------------------------------------"
    print

    if options.plot_map:
        plot_map(foglio[1:], marker='8', markersize=12, color='k')
        plot_map(foglio[1:], marker='8', markersize=11, color='w')
        TPMs=[]
        for i in range(25):
            a=filtro(foglio, 'TPM', i)
            if not a==[]:
                TPMs+=[a]
        for i in range(len(TPMs)):
            plot_map(TPMs[i], marker='o', markersize=12, color=COLORS[i%8])

        a=filtro(foglio, 'Power', "OFF")
        plot_map(a, marker='+', markersize=11, color='k')

        plt.axis([-25,25,-25,25])
        plt.axvline(0, color='b', linestyle='dotted')
        plt.axhline(0, color='b', linestyle='dotted')

        plt.plot([-7.5,7.5],[20,-20],linestyle='dotted', color='b')
        plt.plot([-19,19],[20,-20],linestyle='dotted', color='b')
        plt.plot([-20,20],[8,-8],linestyle='dotted', color='b')
        plt.plot([-20,20],[-7,7],linestyle='dotted', color='b')
        plt.plot([-7.5,7.5],[-20,20],linestyle='dotted', color='b')
        plt.plot([-19,19],[-20,20],linestyle='dotted', color='b')

        plt.annotate("TPM-1", xy=(19,3))
        plt.annotate("TPM-2", xy=(17,11))
        plt.annotate("TPM-3", xy=(10,18))
        plt.annotate("TPM-4", xy=(2,20))
        plt.annotate("TPM-5", xy=(-5,20))
        plt.annotate("TPM-6", xy=(-13,18))
        plt.annotate("TPM-7", xy=(-20,11))
        plt.annotate("TPM-8", xy=(-22,3))
        plt.annotate("TPM-9", xy=(-22,-4))
        plt.annotate("TPM-10", xy=(-20,-12))
        plt.annotate("TPM-11", xy=(-13,-18))
        plt.annotate("TPM-12", xy=(-5,-21))
        plt.annotate("TPM-13", xy=(2,-21))
        plt.annotate("TPM-14", xy=(10,-18))
        plt.annotate("TPM-15", xy=(17,-12))
        plt.annotate("TPM-16", xy=(19,-4))
        #plot_map(foglio[1:], marker='.', markersize=1, color='w', print_name=True)
        plt.show()



