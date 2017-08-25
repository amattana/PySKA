#!/usr/bin/python2.7

# Some globals
COLORS = ['b','g','r','c','m','y','k','w']
RIGHE = 257
COLONNE = 12
EX_FILE = "/home/mattana/Downloads/AAVS 1.1 Locations and connections.xlsx"
OUTPUT_PICTURE_DATA_PATH = "/home/mattana/Documents/AAVS-DATA/"

import urllib3
# Test application, security unimportant:
urllib3.disable_warnings()

import os,sys
from tpm_utils import *
sys.path.append("../board")
sys.path.append("../rf_jig")
sys.path.append("../config")
sys.path.append("../repo_utils")
from bsp.tpm import *
import subprocess
DEVNULL = open(os.devnull,'w')
from ip_scan import *
from optparse import OptionParser
from openpyxl import load_workbook
import matplotlib.pyplot as plt 
import matplotlib.patches as patches
import datetime,time

import gspread
from oauth2client.service_account import ServiceAccountCredentials
import httplib2 

colori=[['        RMS > 30  ','#c800c8'], ['25 < RMS < 30','#ff1d00'], ['20 < RMS < 25','#ff9f00'], ['15 < RMS < 20','#22ff00'], ['10 < RMS < 15','#00ffc5'], ['  5 < RMS < 10 ','#00c5ff'],['    0 < RMS < 5','#0000ff']]

def read_from_google():
    # use creds to create a client to interact with the Google Drive API
    scope = ['https://spreadsheets.google.com/feeds']
    creds = ServiceAccountCredentials.from_json_keyfile_name('client_secret.json', scope)
    client = gspread.authorize(creds)
	 
    # Find a workbook by name and open the first sheet
    # Make sure you use the right name here.
    sheet = client.open("AAVS 1.1 Locations and connections").sheet1


    # Extract and print all of the values
    cells = sheet.get_all_records()
    return cells


def read_from_local():
    if (os.path.isfile(options.EX_FILE)):
        wb2 = load_workbook(options.EX_FILE, data_only=True)
        ws = wb2.active
        wb2.close()
        keys=[]
        for j in range(COLONNE):
            keys += [ws.cell(row=1, column=j+1).value]

        cells=[]
        for i in range(RIGHE-1):
            dic={}
            for j in range(COLONNE):
                val=ws.cell(row=i+2, column=j+1).value
                if not val==None:
                    dic[keys[j]]=val
                else:
                    dic[keys[j]]=""
            cells += [dic]
    else: 
        print "Unable to find file:", options.EX_FILE
        print "\nExiting with errors...\n"
        exit()
    return cells

def print_antennas_list():
    print
    for i in range(nrow):
        print "% 7d\t"%(i),        
        for j in range(ncol):
            #print type(foglio[i][j])
            if type(foglio[i][j])==(unicode or str):
                if len(foglio[i][j])>7:
                    print "%s\t"%(foglio[i][j][:7].center(7,' ')),
                else:
                    print "%s\t"%(foglio[i][j].center(7,' ')),
            elif type(foglio[i][j])==float or type(foglio[i][j])==long:
                if j<9:
                    print "%s\t"%(str("%d"%foglio[i][j]).center(7,' ')),
                else:
                    print "%7.3f\t"%(foglio[i][j]),
            else:
                print " \t",
        print
        if i==0:
            print "--------------------------------------------------------------------------------------------------------"
    print

def plot_map(ant, marker='o', markersize=12, color='g', print_name=False):
    x = [float(str(a['East']).replace(",",".")) for a in ant]
    y = [float(str(a['North']).replace(",",".")) for a in ant]
    name = [a['Base'] for a in ant]
    ax.plot(x,y, marker=marker, markersize=markersize, linestyle = 'None', color=color)
    if print_name:
        for i in range(len(name)):
            ax.annotate("%d"%name[i], xy=(x[i],y[i]), fontsize=10, fontweight='bold')


def onclick(event):                                                        
    if event.dblclick and not event.xdata==None:                                                            
        if event.button == 1:
            sel=[x for x in cells if ((x['East']>event.xdata-0.4) and (x['East']<event.xdata+0.4))] 
            res=[x for x in sel   if ((x['North']>event.ydata-0.4) and (x['North']<event.ydata+0.4))]
            if len(res)==1:
                board = {}
                print "Selected antenna", int(res[0]['Base'])#, len(TPMs)#,res[0]['East'], res[0]['North']
                for i in range(len(TPMs)):
                    #for x in TPMs[i]['ANTENNE']:
                    #    print x['Base'], " ",
                    #print
                    if len([x for x in TPMs[i]['ANTENNE'] if int(x['Base'])==int(res[0]['Base'])])>0:
                        board['IP'] = TPMs[i]['IP']
                        board['TPM'] = TPMs[i]['TPM']
                        board['ANTENNE'] = [x for x in TPMs[i]['ANTENNE'] if int(x['Base'])==int(res[0]['Base'])]
                        #print board['ANTENNE'], board['IP'] 
                if not board=={}:
                    if len(board['ANTENNE'])==1:
                        freqs, spettro=get_raw_meas(board, meas="SPECTRA")
                        fig_spectrum=plt.figure(num=2, figsize=(12,9), dpi=80, facecolor='w', edgecolor='w')
                        axs = fig_spectrum.add_axes([0.08, 0.13, 0.85, 0.75])
                        axs.axis([0,400,0,-100])
                        axs.set_ylim([-100, 0])
                        axs.set_xlim([0, 400])
                        axs.set_xlabel('MHz', fontsize=14)
                        axs.set_ylabel('dB', fontsize=14)
                        msg = "Antenna Base # "+str(int(board['ANTENNE'][0]['Base']))+"     "
                        msg += "Hybrid Cable: "+str(int(board['ANTENNE'][0]['Hybrid Cable']))+"     "
                        msg += "Roxtec: "+str(int(board['ANTENNE'][0]['Roxtec']))+"\n"
                        msg += "Ribbon: "+str(int(board['ANTENNE'][0]['Ribbon']))+"    "
                        msg += "Fibre: "+str(int(board['ANTENNE'][0]['Fibre']))+"    "
                        msg += "Colour: "+str(board['ANTENNE'][0]['Colour'])+"\n"
                        msg += "TPM: "+str(int(board['ANTENNE'][0]['TPM']))+"   IP:"+board['IP']+"    "
                        msg += "RX: "+str(int(board['ANTENNE'][0]['RX']))
                        fig_spectrum.suptitle(msg, fontsize=16)
                        axs.plot(freqs, spettro[(int(board['ANTENNE'][0]['RX'])-1)*2], scaley=False, color='b', label="Pol X")
                        axs.plot(freqs, spettro[(int(board['ANTENNE'][0]['RX'])-1)*2+1], scaley=False, color='g', label="Pol Y")
                        axs.legend(bbox_to_anchor=(0.36, -0.16, 1., .102), loc=3,ncol=2)#, mode="expand", borderaxespad=0.)
                        fig_spectrum.show()

            elif len(res)>2:
                print "Search provides more than one result (found %d candidates)"%(len(res))
            else:
                #print "Double clicked on x:%4.2f and y:%4.2f, no antenna found here!"%(event.xdata,event.ydata)
                pass
        else:                                                            
            pass                                                                                            
    else:
        pass


if __name__ == "__main__":
    parser = OptionParser()
    parser.add_option("-e", "--excel_file", 
                      dest="EX_FILE", 
                      default = "/home/mattana/Downloads/AAVS 1.1 Locations and connections.xlsx",
                      help="The excel file containing the AAVS 1.1 SpreadSheet")
                      
    parser.add_option("-p", "--noplot", action='store_true',
                      dest="no_plot", 
                      default = False,
                      help="If True do not plot the figure")
                      
    parser.add_option("-s", "--nosave", action='store_true',
                      dest="no_save", 
                      default = False,
                      help="If True do not save the figure")
                      
    parser.add_option("-t", "--notpm", action='store_true',
                      dest="no_tpm", 
                      default = False,
                      help="If True do not list TPMs in the figure")
                      
    parser.add_option("-c", "--nocoord", action='store_true',
                      dest="no_coord", 
                      default = False,
                      help="If True do not show coords in the figure")
                      
    parser.add_option("-f", "--nooff", action='store_true',
                      dest="no_off", 
                      default = False,
                      help="If True do not show OFF antenna in the figure")
                      
    parser.add_option("-o", "--noon", action='store_true',
                      dest="no_on", 
                      default = False,
                      help="If True do not colour the antenna ON in the figure")
                      
    parser.add_option("-a", "--noall", action='store_true',
                      dest="no_all", 
                      default = False,
                      help="If True do not show empty bases in the figures")
                      
    parser.add_option("-n", "--name", action='store_true',
                      dest="ant_name", 
                      default = False,
                      help="If True show antenna bases numerber in the figures")
                      
    parser.add_option("-l", "--list", action='store_true',
                      dest="list_ant", 
                      default = False,
                      help="If enabled list the antenna map")
                      
    parser.add_option("--local", action='store_true',
                      dest="force_local",
                      default = False,
                      help="If enabled force the program to run with the locally stored excel file")

    parser.add_option("--meas", 
                      dest="meas",
                      default = "",
                      help="Choose from RMS (ADU RMS) or DBM (RF Power in dBm)")

    parser.add_option("--plot_pol", 
                      dest="plot_pol",
                      default = "x",
                      help="Select the polarization [x|y] to be shown")


    (options, args) = parser.parse_args()


    if not options.force_local:
        try: 
            cells = read_from_google()
            print "\nSuccessfully connected to the online google spreadsheet!\n\n"

        except httplib2.ServerNotFoundError:
            print("\nUnable to find the server at accounts.google.com.\n\nContinuing with localfile: %s\n"%(EX_FILE))
            cells = read_from_local()

        except:
            print "Got a new exception! Exiting..."
            exit()
    else:
        print "\nRunning with local file:", options.EX_FILE
        cells = read_from_local()

    fig=plt.figure(num=None, figsize=(12,9), dpi=80, facecolor='w', edgecolor='w')

    ax = fig.add_axes([0.08, 0.08, 0.7, 0.85])
    ax.set_title("AAVS 1.1 Antennas Map")

    ax.axis([-25,25,-25,25])

    ax.axvline(0, color='b', linestyle='dotted')
    ax.axhline(0, color='b', linestyle='dotted')

    if not options.no_tpm:
        ax.plot([-7.5,7.5],[20,-20],linestyle='dotted', color='b')
        ax.plot([-19,19],[20,-20],linestyle='dotted', color='b')
        ax.plot([-20,20],[8,-8],linestyle='dotted', color='b')
        ax.plot([-20,20],[-7,7],linestyle='dotted', color='b')
        ax.plot([-7.5,7.5],[-20,20],linestyle='dotted', color='b')
        ax.plot([-19,19],[-20,20],linestyle='dotted', color='b')

        ax.annotate("TPM-1", xy=(19,3))
        ax.annotate("TPM-2", xy=(17,11))
        ax.annotate("TPM-3", xy=(10,18))
        ax.annotate("TPM-4", xy=(2,20))
        ax.annotate("TPM-5", xy=(-5,20))
        ax.annotate("TPM-6", xy=(-13,18))
        ax.annotate("TPM-7", xy=(-20,11))
        ax.annotate("TPM-8", xy=(-22,3))
        ax.annotate("TPM-9", xy=(-22,-4))
        ax.annotate("TPM-10", xy=(-20,-12))
        ax.annotate("TPM-11", xy=(-13,-18))
        ax.annotate("TPM-12", xy=(-5,-21))
        ax.annotate("TPM-13", xy=(2,-21))
        ax.annotate("TPM-14", xy=(10,-18))
        ax.annotate("TPM-15", xy=(17,-12))
        ax.annotate("TPM-16", xy=(19,-4))

    if not options.no_coord:
        ax.annotate("NORTH", xy=(-2,22), fontweight='bold')
        ax.annotate("SOUTH", xy=(-1.9,-24), fontweight='bold')
        ax.annotate("EAST", xy=(21.5,0), fontweight='bold')
        ax.annotate("WEST", xy=(-24.5,0), fontweight='bold')

    if not options.no_all:
        plot_map(cells, marker='8', markersize=12, color='k')
        plot_map(cells, marker='8', markersize=11, color='w', print_name=options.ant_name)

    tpms=ip_scan()

    TPMs=[]
    for i in tpms:
        tpm = {}
        tpm['TPM'] = TPM(ip=i, port=10000, timeout=1)
        tpm['IP']  = i
        tpm['ANTENNE'] = [x for x in cells if x['TPM']==int(i.split(".")[-1])]
        TPMs += [tpm]

    if not options.meas=="":
        if options.plot_pol=='x':
            pol=0
            #ax.add_patch(patches.Rectangle((20, 20),    # (x,y)
            #                                 8,         # width
            #                                 2,         # height
            #                                facecolor="red"))

            ax.text(30, 20, "-       Pol X       -", bbox={'facecolor': '#22ff00', 'pad': 10})
            ax.text(30, 17, "-       Pol Y       -", bbox={'facecolor': '#ff1d00', 'pad': 10})
        else: 
            pol=1
        for tpm in TPMs:
            rms=get_raw_meas(tpm, meas=options.meas)
            for j in range(len(rms)/2):
                x=float(str(tpm['ANTENNE'][j]['East']).replace(",","."))
                y=float(str(tpm['ANTENNE'][j]['North']).replace(",","."))
                ax.plot(x,y, marker='8', markersize=10, linestyle = 'None', color=rms_color(rms[(j*2)+pol]))

        for i in range(len(colori)):
            ax.text(30, -6-i*3, colori[i][0], bbox={'facecolor': colori[i][1], 'pad': 10})
    print
    pressed_x=0
    pressed_y=0
    fig.canvas.mpl_connect('button_press_event', onclick)
    
    plt.show()
    exit()

    if not options.no_on:
        for i in range(len(TPMs)):
            plot_map(TPMs[i], marker='8', markersize=10, color=COLORS[i%8])

    if not options.no_off:
        spente=[a for a in cells if "OFF" in a['Power']]
        plot_map(spente, marker='+', markersize=11, color='k')


    if not options.no_save:
        fname=OUTPUT_PICTURE_DATA_PATH+datetime.datetime.strftime(datetime.datetime.utcfromtimestamp(time.time()),"%Y-%m-%d_%H%M%S_AAVS_MAP.png")
        fig.savefig(fname)
        print "Saved picture", fname
    if not options.no_plot:
        plt.show()

    if options.list_ant:
        print_antennas_list()

