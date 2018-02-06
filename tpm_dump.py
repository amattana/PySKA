#!/usr/bin/env python

'''

  AAVS Logger.

  It produces for each antenna and for both pols:
    -  Time domain binary data (first double word (64b) is the lenght of the following double word (64b) elements)
    -  Spectra binary data (first double word (64b) is the lenght of the following double word (64b) elements)
    -  Picture of the spectra

  Logging period can be specified in minutes with parameter -t (--time)

  When hit Ctrl+C (Keyboard Interrupt Signal) it produces
    -  A Movie (MPEG4 avi) for each antenna saved in the videos folder with subfolders for each pol

'''

__author__ = "Andrea Mattana"
__copyright__ = "Copyright 2017, Istituto di RadioAstronomia, INAF, Italy"
__credits__ = ["Andrea Mattana"]
__license__ = "GPL"
__version__ = "1.0"
__maintainer__ = "Andrea Mattana"

import sys
sys.path.append("../board")
sys.path.append("../rf_jig")
sys.path.append("../config")
sys.path.append("../repo_utils")
from tpm_utils import *
from bsp.tpm import *
DEVNULL = open(os.devnull,'w')

from gui_utils import *
from rf_jig import *
from rfjig_bsp import *
from ip_scan import *
import subprocess

import gspread
from oauth2client.service_account import ServiceAccountCredentials
import httplib2
from openpyxl import load_workbook

from optparse import OptionParser

import matplotlib.pyplot as plt

import urllib3
# Test application, security unimportant:
urllib3.disable_warnings()


# Other stuff
import numpy as np
import struct
import datetime

# Some globals
COLONNE = 20
RIGHE = 256
COLORS = ['b','g','r','c','m','y','k','w']
<<<<<<< HEAD
OUT_PATH = "./data/"
=======
OUT_PATH = "/data/data_2/2017-11-AAVS-FIX/"
>>>>>>> 5001e6e33527d3c355c9d63d5e75999e59c0edf6
MAP_PATH = "Maps"
MOVIE_FOLDER = "Videos"
PATH_PLOT_LIST = "./.plotlists/"
EX_FILE = "/home/mattana/Downloads/AAVS 1.1 Locations and connections.xlsx"
EX_FILE_AAVS = "/home/aavs/Downloads/AAVS 1.1 Locations and connections.xlsx"

def read_from_google():
    # use creds to create a client to interact with the Google Drive API
    scope = ['https://spreadsheets.google.com/feeds']
    creds = ServiceAccountCredentials.from_json_keyfile_name('client_secret.json', scope)
    client = gspread.authorize(creds)

    # Find a workbook by name and open the first sheet
    # Make sure you use the right name here.
    sheet = client.open("AAVS 1.1 Locations and connections").sheet1


    # Extract and print all of the values
    cells = sheet.get_all_records()
    for i in range(len(cells)):
        cells[i]['East'] = float(cells[i]['East'].replace(",","."))
        cells[i]['North'] = float(cells[i]['North'].replace(",", "."))
    return cells


def read_from_local(fname):
    if (os.path.isfile(fname)):
        wb2 = load_workbook(fname, data_only=True)
        ws = wb2.active
        wb2.close()
        keys=[]
        for j in range(COLONNE):
            keys += [ws.cell(row=1, column=j+1).value]

        cells=[]
        for i in range(RIGHE-1):
            dic={}
            for j in range(COLONNE):
                val=ws.cell(row=i+2, column=j+1).value
                if not val==None:
                    dic[keys[j]]=val
                else:
                    dic[keys[j]]=""
            cells += [dic]
    else:
        print "Unable to find file:", fname
        print "\nExiting with errors...\n"
        exit()
    return cells


def loadAAVSdata(debug=False):
    if not debug:
        try:
            cells = read_from_google()
            print "\nSuccessfully connected to the online google spreadsheet!\n\n"

        except httplib2.ServerNotFoundError:
            print("\nUnable to find the server at accounts.google.com.\n\nContinuing with local file: %s\n" % (EX_FILE))
            cells = read_from_local()
            print "done!"
    else:
        if os.path.isfile(EX_FILE):
            print("\nReading local file: %s\n" % (EX_FILE))
            cells = read_from_local(EX_FILE)
        else:
            print("\nReading local file: %s\n" % (EX_FILE_AAVS))
            cells = read_from_local(EX_FILE_AAVS)
    return cells

def plotta(dati,fname,pol):
    if pol=="X" or pol==0:
        plt.plot(np.linspace(0,400,len(dati[1:])),dati[1:],color='b')
    else:
        plt.plot(np.linspace(0,400,len(dati[1:])),dati[1:],color='g')
    plt.xlim(0,400)
    plt.ylim(-80,-20)
    plt.title(fname[fname.rfind("/")+1:])
    plt.savefig(fname+".png")
    plt.clf()


def plot_map(ant, marker='o', markersize=12, color='g', print_name=False):
    x = [float(str(a['East']).replace(",", ".")) for a in ant]
    y = [float(str(a['North']).replace(",", ".")) for a in ant]
    # name = [a['Hybrid Cable'] for a in ant]
    name = [a['Base'] for a in ant]
    ax.plot(x, y, marker=marker, markersize=markersize, linestyle='None', color=color)
    if print_name:
        for i in range(len(name)):
            # ax.annotate("%s"%name[i], xy=(x[i],y[i]), fontsize=10, fontweight='bold')
            ax.annotate("%d" % name[i], xy=(x[i], y[i]), fontsize=10, fontweight='bold')



def generateMap(cells, tpm_used, rms_map, timestamp, outdir, mapdir):
    dataora = datetime.datetime.strptime(timestamp,"%Y-%m-%d_%H%M%S")
    for pol in enumerate(["X","Y"]):
        fig=plt.figure(num=1, figsize=(12,9), dpi=80, facecolor='w', edgecolor='w')
        ax = fig.add_axes([0.08, 0.08, 0.7, 0.85])
        ax.set_title("AAVS 1.1 Antennas Map - " + pol[1] + " Pol")

        ax.axis([-25, 25, -25, 25])

        ax.axvline(0, color='b', linestyle='dotted')
        ax.axhline(0, color='b', linestyle='dotted')

        ax.plot([-7.5, 7.5], [20, -20], linestyle='dotted', color='b')
        ax.plot([-19, 19], [20, -20], linestyle='dotted', color='b')
        ax.plot([-20, 20], [8, -8], linestyle='dotted', color='b')
        ax.plot([-20, 20], [-7, 7], linestyle='dotted', color='b')
        ax.plot([-7.5, 7.5], [-20, 20], linestyle='dotted', color='b')
        ax.plot([-19, 19], [-20, 20], linestyle='dotted', color='b')

        ax.annotate("TPM-1", xy=(19, 3), fontsize=10)
        ax.annotate("TPM-2", xy=(17, 11), fontsize=10)
        ax.annotate("TPM-3", xy=(10, 18), fontsize=10)
        ax.annotate("TPM-4", xy=(2, 20), fontsize=10)
        ax.annotate("TPM-5", xy=(-5, 20), fontsize=10)
        ax.annotate("TPM-6", xy=(-13, 18), fontsize=10)
        ax.annotate("TPM-7", xy=(-20, 11), fontsize=10)
        ax.annotate("TPM-8", xy=(-22, 3), fontsize=10)
        ax.annotate("TPM-9", xy=(-22, -4), fontsize=10)
        ax.annotate("TPM-15", xy=(-20, -12), fontsize=10)
        ax.annotate("TPM-14", xy=(-13, -18), fontsize=10)
        ax.annotate("TPM-13", xy=(-5, -21), fontsize=10)
        ax.annotate("TPM-12", xy=(2, -21), fontsize=10)
        ax.annotate("TPM-11", xy=(10, -18), fontsize=10)
        ax.annotate("TPM-10", xy=(17, -12), fontsize=10)
        ax.annotate("TPM-9", xy=(19, -4), fontsize=10)

        ax.annotate("NORTH", xy=(-2, 22), fontweight='bold', fontsize=9)
        ax.annotate("SOUTH", xy=(-1.9, -24), fontweight='bold', fontsize=9)
        ax.annotate("EAST", xy=(21.5, 0), fontweight='bold', fontsize=9)
        ax.annotate("WEST", xy=(-24.5, 0), fontweight='bold', fontsize=9)
        ax.text(-8, -28.5, datetime.datetime.strftime(dataora, "%Y/%m/%d %H:%M:%S UTC"), fontsize=16)
        print len(rms_map)
        for tpm in range(1, tpm_used):
            #print tpm, " * "
            rms = [a for a in rms_map if a[0]=="10.0.10."+str(tpm)]
            if rms == []:
                continue
            rms = rms[0]
            for rx in range(1, 16 + 1):
                cella = [a for a in cells if ((a['TPM'] == tpm) and (a['RX'] == rx))]
                if not cella==[]:
                    x = cella[0]['East']
                    y = cella[0]['North']
                    ax.plot(x,y, marker='8', markersize=10, linestyle = 'None', color=rms_color(rms[1][((rx-1)*2) + pol[0]]))
        plt.draw()
        plt.savefig(outdir + mapdir + "/Pol-" + pol[1] + "/AAVS1_MAP_Pol-" + pol[1] + timestamp + ".png")
        plt.close()

if __name__ == "__main__":
    parser = OptionParser()
    parser.add_option("-t", "--time",
                      dest="time",
                      default=5,
                      help="Time interval in minutes")

<<<<<<< HEAD
    parser.add_option("-b", "--board",
                      dest="board",
                      default=1,
                      help="Board number # (10.0.10.#)")


=======
>>>>>>> 5001e6e33527d3c355c9d63d5e75999e59c0edf6

    (options, args) = parser.parse_args()

    print "\n######################################################"
<<<<<<< HEAD
    print "\n\nTPM Data Logger"
    print "\n\n  - Time interval is "+str(options.time)+"minutes"
    #print "\n\nReading AAVS1 Google Spreadsheet"

    # Read google Spreadsheet
    #cells = loadAAVSdata()
    #tpm_used = max([a['TPM'] for a in cells if not a['TPM'] == ""])+1
    tpm_used=1

    # Search for TPMs
    TPMs = ["10.0.10."+str(options.board)]
=======
    print "\n\nAAVS1 Data Logger"
    print "\n\n  - Time interval is "+str(options.time)+"minutes"
    print "\n\nReading AAVS1 Google Spreadsheet"

    # Read google Spreadsheet
    cells = loadAAVSdata()
    tpm_used = max([a['TPM'] for a in cells if not a['TPM'] == ""])+1
    tpm_used=16

    # Search for TPMs
    TPMs = ip_scan()
>>>>>>> 5001e6e33527d3c355c9d63d5e75999e59c0edf6
    data = datetime.datetime.strftime(datetime.datetime.utcfromtimestamp(time.time()),"%Y/%m/%d")
    ora = datetime.datetime.strftime(datetime.datetime.utcfromtimestamp(time.time()),"%H:%M:%S")
    print "\nStart Logging on " + data + " at " + ora + "\n\n"
    #TPMs = ['10.0.10.{0}'.format(i+1) for i in xrange(16)]

    # Creating Main data Directory
    if not os.path.exists(OUT_PATH):
        os.makedirs(OUT_PATH)
        print "\nGenerating main directory for data... (" + OUT_PATH + ")"

    # Creating Directory for today's data
    OUT_PATH = OUT_PATH + datetime.datetime.strftime(datetime.datetime.utcfromtimestamp(time.time()),"%Y-%m-%d/")
    if not os.path.exists(OUT_PATH):
        os.makedirs(OUT_PATH)
        print "\nGenerating directory for today data... (" + OUT_PATH + ")"

<<<<<<< HEAD
    ## Creating Directory to store the videos
    #if not os.path.exists(OUT_PATH + MOVIE_FOLDER):
    #    os.makedirs(OUT_PATH + MOVIE_FOLDER)
    #    os.makedirs(OUT_PATH + MOVIE_FOLDER + "/Pol-X")
    #    os.makedirs(OUT_PATH + MOVIE_FOLDER + "/Pol-Y")
    #    os.makedirs(OUT_PATH + MAP_PATH)
    #    os.makedirs(OUT_PATH + MAP_PATH + "/Pol-X")
    #    os.makedirs(OUT_PATH + MAP_PATH + "/Pol-Y")
    #    print "\nGenerating directories for Videos... (" + OUT_PATH + MOVIE_FOLDER + ") + Pol-X and Pol-Y"

    plt.ioff()
    counter = 0
    try:
        while True:
            counter = counter +1
=======
    # Creating Directory to store the videos
    if not os.path.exists(OUT_PATH + MOVIE_FOLDER):
        os.makedirs(OUT_PATH + MOVIE_FOLDER)
        os.makedirs(OUT_PATH + MOVIE_FOLDER + "/Pol-X")
        os.makedirs(OUT_PATH + MOVIE_FOLDER + "/Pol-Y")
        os.makedirs(OUT_PATH + MAP_PATH)
        os.makedirs(OUT_PATH + MAP_PATH + "/Pol-X")
        os.makedirs(OUT_PATH + MAP_PATH + "/Pol-Y")
        print "\nGenerating directories for Videos... (" + OUT_PATH + MOVIE_FOLDER + ") + Pol-X and Pol-Y"

    plt.ioff()
    try:
        while True:
>>>>>>> 5001e6e33527d3c355c9d63d5e75999e59c0edf6
            ora = datetime.datetime.strftime(datetime.datetime.utcfromtimestamp(time.time()),"%Y-%m-%d_%H%M%S")
            rms_map = []
            for i in  TPMs:
                tpm = int(i.split(".")[-1])
<<<<<<< HEAD
                #check=[x for x in cells if (x['TPM']==int(i.split(".")[-1]))]
                #if check==[]:
                #    continue
=======
                check=[x for x in cells if (x['TPM']==int(i.split(".")[-1]))]
                if check==[]:
                    continue
>>>>>>> 5001e6e33527d3c355c9d63d5e75999e59c0edf6
                if not os.path.exists(OUT_PATH+i):
                    os.makedirs(OUT_PATH+i)
                freqs, spettro, rawdata , rms, rfpower = get_raw_meas(tpm_obj(i))
                rms_map += [[i,rms]]
                #print len(spettro)
                for rx in xrange(len(spettro)/2):
                    for p in xrange(2):
                        fname = OUT_PATH + i + "/"
                        fname += str("RX-%02d/" % (int(rx + 1)))
                        if not os.path.exists(fname):
                            os.makedirs(fname)
                        if p%2==0:
                            pol="X"
                            fname += "Pol-X/"
                            if not os.path.exists(fname):
                                os.makedirs(fname)
<<<<<<< HEAD
                            fname += str("TPM-%02d" % (int(i.split(".")[-1]))) + str("_RX-%02d" % (int(rx + 1))) #+ "_BASE-"
                            #base = [x for x in cells if (x['TPM'] == int(i.split(".")[-1]) and x['RX'] == (rx + 1))]
                            #if base == []:
                            #    continue
                            #fname += str("%03d" % (base[0]['Base']))
=======
                            fname += str("TPM-%02d" % (int(i.split(".")[-1]))) + str("_RX-%02d" % (int(rx + 1))) + "_BASE-"
                            base = [x for x in cells if (x['TPM'] == int(i.split(".")[-1]) and x['RX'] == (rx + 1))]
                            if base == []:
                                continue
                            fname += str("%03d" % (base[0]['Base']))
>>>>>>> 5001e6e33527d3c355c9d63d5e75999e59c0edf6
                            fname += "_Pol-X_" + ora
                        else:
                            pol="Y"
                            fname += "Pol-Y/"
                            if not os.path.exists(fname):
                                os.makedirs(fname)
<<<<<<< HEAD
                            fname += str("TPM-%02d"%(int(i.split(".")[-1])))+str("_RX-%02d"%(int(rx+1)))#+"_BASE-"
                            #base = [x for x in cells if (x['TPM']==int(i.split(".")[-1]) and x['RX']==(rx+1))]
                            #if base==[]:
                            #    continue
                            #fname += str("%03d"%(base[0]['Base']))
                            fname += "_Pol-Y_" + ora

                        #if not base==[]:
                        with open(fname+".spe","wb") as f:
                            f.write(struct.pack(">d", len(spettro[(rx*2)+p])))
                            f.write(struct.pack(">"+str(len(spettro[(rx*2)+p]))+"d", *spettro[(rx*2)+p]))
                        with open(fname+".tdd","wb") as f:
                            f.write(struct.pack(">d", len(rawdata[(rx*2)+p])))
                            f.write(struct.pack(">"+str(len(rawdata[(rx*2)+p]))+"d", *rawdata[(rx*2)+p]))
                        plotta(spettro[(rx*2)+p],fname,pol)

                        print "Written files:",fname
                print
            #print "\nGenerating AAVS1 MAP for samples"+ora+"... ",
            #generateMap(cells, tpm_used, rms_map, ora, OUT_PATH, MAP_PATH)
 
            print "\n\nMeasurements #",counter, "acquired!\n\n"
            print datetime.datetime.strftime(datetime.datetime.utcfromtimestamp(time.time()),"%Y/%m/%d %H:%M:%S"),
            print "  -  Next acquisition will be done in ",options.time," minutes"
            time.sleep(60*int(options.time))
            print "\n",datetime.datetime.strftime(datetime.datetime.utcfromtimestamp(time.time()),"%Y/%m/%d %H:%M:%S"),
            print "  -  Starting acquisition #",counter+1,"\n"
    except KeyboardInterrupt:
        print "\n\nProgram terminated!\n\n"

    #    for tpm in TPMs:
    #        for rx in range(1,17):
    #            for pol in ["X","Y"]:
    #                data_path = OUT_PATH + tpm + "/RX-" + str("%02d" % (rx)) + "/Pol-" + pol
    #                if os.path.exists(data_path):
    #                    fname = "TPM-" + str("%02d" % (int(tpm.split(".")[-1]))) + "_RX-" + str("%02d" % (rx)) + "_Pol-" + pol + ".avi"
    #                    print "\n\nGenerating movie for " + fname
    #                    cmd = "ffmpeg -f image2 -i " + data_path + "/%*.png  " + OUT_PATH + MOVIE_FOLDER + "/Pol-" + pol + "/" + fname
    #                    print "\n",cmd,"\n\n"
    #                    os.system(cmd)

# from matplotlib import pyplot as plt
# import struct
#
# with open("/home/mattana/aavs_data/2018-02-05/10.0.10.1/RX-01/Pol-X/TPM-01_RX-01_Pol-X_2018-02-05_164804.spe","r") as f:
#     a=f.read()
# l = struct.unpack(">d",a[0:8])[0]
# data=struct.unpack(">"+str(int(l))+"d",a[8:])
# plt.plot(data)
# plt.show()
=======
                            fname += str("TPM-%02d"%(int(i.split(".")[-1])))+str("_RX-%02d"%(int(rx+1)))+"_BASE-"
                            base = [x for x in cells if (x['TPM']==int(i.split(".")[-1]) and x['RX']==(rx+1))]
                            if base==[]:
                                continue
                            fname += str("%03d"%(base[0]['Base']))
                            fname += "_Pol-Y_" + ora

                        if not base==[]:
                            with open(fname+".spe","wb") as f:
                                f.write(struct.pack(">d", len(spettro[(rx*2)+p])))
                                f.write(struct.pack(">"+str(len(spettro[(rx*2)+p]))+"d", *spettro[(rx*2)+p]))
                            with open(fname+".tdd","wb") as f:
                                f.write(struct.pack(">d", len(rawdata[(rx*2)+p])))
                                f.write(struct.pack(">"+str(len(rawdata[(rx*2)+p]))+"d", *rawdata[(rx*2)+p]))
                            plotta(spettro[(rx*2)+p],fname,pol)

                            print "Written files:",fname
                print
            print "\nGenerating AAVS1 MAP for samples"+ora+"... ",
            generateMap(cells, tpm_used, rms_map, ora, OUT_PATH, MAP_PATH)

            print datetime.datetime.strftime(datetime.datetime.utcfromtimestamp(time.time()),"%Y/%m/%d %H:%M:%S"),
            print "\n  -  Next acquisition will be done in ",options.time," minutes"
            time.sleep(60*int(options.time))
            print "\n",datetime.datetime.strftime(datetime.datetime.utcfromtimestamp(time.time()),"%Y/%m/%d %H:%M:%S"),
            print "  -  Starting new acquisition\n"
    except KeyboardInterrupt:

        for tpm in TPMs:
            for rx in range(1,17):
                for pol in ["X","Y"]:
                    data_path = OUT_PATH + tpm + "/RX-" + str("%02d" % (rx)) + "/Pol-" + pol
                    if os.path.exists(data_path):
                        fname = "TPM-" + str("%02d" % (int(tpm.split(".")[-1]))) + "_RX-" + str("%02d" % (rx)) + "_Pol-" + pol + ".avi"
                        print "\n\nGenerating movie for " + fname
                        cmd = "ffmpeg -f image2 -i " + data_path + "/%*.png  " + OUT_PATH + MOVIE_FOLDER + "/Pol-" + pol + "/" + fname
                        print "\n",cmd,"\n\n"
                        os.system(cmd)


# with open("/home/mattana/aavs_data/10.0.10.3/2017-11-07_160741_TPM-03_RX-16_BASE-54.0_Pol-Y.dat","r") as f:
#     a=f.read()
# data=struct.unpack(">4097d",a)
# plt.plot(data)
>>>>>>> 5001e6e33527d3c355c9d63d5e75999e59c0edf6

