#!/usr/bin/env python

'''

  GUI App for debugging AAVS1 antennas during installation on the MRO field

'''

__author__ = "Andrea Mattana"
__copyright__ = "Copyright 2017, Istituto di RadioAstronomia, INAF, Italy"
__credits__ = ["Andrea Mattana"]
__license__ = "GPL"
__version__ = "1.0"
__maintainer__ = "Andrea Mattana"


from PyQt4 import QtCore, QtGui, uic
import sys, os, socket
sys.path.append("../board")
sys.path.append("../rf_jig")
sys.path.append("../config")
sys.path.append("../repo_utils")
#sys.path.append("../board/pyska")
from tpm_utils import *
#from ska_tpm import xmlparser
#from jig_adu_test import *
from bsp.tpm import *
#import config.manager as config_man
import manager as config_man
from netproto.sdp_medicina import sdp_medicina as sdp_med
import subprocess
DEVNULL = open(os.devnull,'w')

import datetime, time
import sys, easygui, datetime
#from qt_rf_jig_utils import *
from gui_utils import *
from rf_jig import *
from rfjig_bsp import *
from ip_scan import *

import gspread
from oauth2client.service_account import ServiceAccountCredentials
import httplib2
from openpyxl import load_workbook

from optparse import OptionParser

# Matplotlib stuff
import matplotlib.patches as mpatches
from matplotlib.backends.backend_qt4agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.pyplot as plt


# Other stuff
import numpy as np
import struct

#import multiprocessing
from threading import Thread

# Some globals

COLORS = ['b','g','r','k','y','c']
RIGHE = 257
COLONNE = 24
EX_FILE = "/home/mattana/Downloads/AAVS 1.1 Locations and connections.xlsx"
EX_FILE_AAVS = "/home/aavs/Downloads/AAVS 1.1 Locations and connections.xlsx"
OUTPUT_PICTURE_DATA_PATH = "/home/mattana/Documents/AAVS-DATA/"
LOG_PATH = "/home/mattana/aavs_data"
PATH_PLOT_LIST = "./.plotlists/"
LABEL_WIDTH  = 23
LABEL_HEIGHT = 15
TEXT_WIDTH   = 50
TEXT_HEIGHT  = 22
FLAG_WIDTH   = 25
FLAG_HEIGHT  = 16

TABLE_HSPACE = 430
TABLE_VSPACE = 30

import urllib3
# Test application, security unimportant:
urllib3.disable_warnings()

try:
    _fromUtf8 = QtCore.QString.fromUtf8
except AttributeError:
    def _fromUtf8(s):
        return s

try:
    _encoding = QtGui.QApplication.UnicodeUTF8
    def _translate(context, text, disambig):
        return QtGui.QApplication.translate(context, text, disambig, _encoding)
except AttributeError:
    def _translate(context, text, disambig):
        return QtGui.QApplication.translate(context, text, disambig)

colori=[['        RMS > 30  ','#c800c8'], ['25 < RMS < 30','#ff1d00'], ['20 < RMS < 25','#ff9f00'], ['15 < RMS < 20','#22ff00'], ['10 < RMS < 15','#00ffc5'], ['  5 < RMS < 10 ','#00c5ff'],['    0 < RMS < 5','#0000ff']]
colori_tpm=['b','g','r','y']

map_tpm_range = [['TPM-1',[18.8,4.1,22.3,2.8],'b'],
             ['TPM-2',[16.57,12,20.24,10.77],'g'],
             ['TPM-3',[9.14,19,13.58,17.81],'r'],
             ['TPM-4',[1.45,21.37,5.29,19.4],'y'],
             ['TPM-5',[-5.97,21.28,-1.62,19.68],'b'],
             ['TPM-6',[-14.09,19.59,-8.54,17.52],'g'],
             ['TPM-7',[-21.35,12.64,-16.31,10.49],'r'],
             ['TPM-8',[-22.89,4.48,-18.62,2.41],'y'],
             ['TPM-9',[18.45,-2.64,22.55,-4.52],'y'],
             ['TPM-10',[16.14,-10.43,21.18,-12.69],'r'],
             ['TPM-11',[9.31,-16.25,14,-18.7],'g'],
             ['TPM-12',[1,-19.35,6.23,-21.7],'b'],
             ['TPM-13',[-5.97,-19.34,-1.1,-21.7],'y'],
             ['TPM-14',[-13.75,-16.16,-8.9,-18.8],'r'],
             ['TPM-15',[-20.84,-10.34,-16.05,-12.78],'g'],
             ['TPM-16',[-22.8,-2.27,-18,-4.6],'b'],
            ]

def read_from_google():
    # use creds to create a client to interact with the Google Drive API
    scope = ['https://spreadsheets.google.com/feeds']
    creds = ServiceAccountCredentials.from_json_keyfile_name('client_secret.json', scope)
    client = gspread.authorize(creds)

    # Find a workbook by name and open the first sheet
    # Make sure you use the right name here.
    sheet = client.open("AAVS 1.1 Locations and connections").sheet1


    # Extract and print all of the values
    cells = sheet.get_all_records()
    for i in range(len(cells)):
        cells[i]['East'] = float(cells[i]['East'].replace(",","."))
        cells[i]['North'] = float(cells[i]['North'].replace(",", "."))
    return cells


def read_from_local(fname):
    if (os.path.isfile(fname)):
        wb2 = load_workbook(fname, data_only=True)
        ws = wb2.active
        wb2.close()
        keys=[]
        for j in range(COLONNE):
            keys += [ws.cell(row=1, column=j+1).value]

        cells=[]
        for i in range(RIGHE-1):
            dic={}
            for j in range(COLONNE):
                val=ws.cell(row=i+2, column=j+1).value
                if not val==None:
                    dic[keys[j]]=val
                else:
                    dic[keys[j]]=""
            cells += [dic]
    else:
        print "Unable to find file:", fname
        print "\nExiting with errors...\n"
        exit()
    return cells


# This creates the input label (eg: for input 15 -> "15:")
def create_label(Dialog, x, y, w, h, text):
    label = QtGui.QLabel(Dialog)
    label.setGeometry(QtCore.QRect(x, y, w, h))
    label.setAlignment(QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
    label.setText(_translate("Dialog", text, None))
    label.setFont(QtGui.QFont("Ubuntu",9))
    return label

def create_flag(Dialog, x, y, color, text):
    flag = QtGui.QLabel(Dialog)
    flag.setGeometry(QtCore.QRect(x, y, FLAG_WIDTH, FLAG_HEIGHT))
    flag.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
    flag.setAutoFillBackground(True)
    if color=="green":
        flag.setStyleSheet(_fromUtf8("background-color: rgb(0, 170, 0);"))
    elif color=="yellow":
        flag.setStyleSheet(_fromUtf8("background-color: rgb(255, 255, 0);"))
    elif color=="cyan":
        flag.setStyleSheet(_fromUtf8("background-color: rgb(0, 255, 234);"))
        flag.setCursor(QtGui.QCursor(QtCore.Qt.ArrowCursor))
    else:
        flag.setStyleSheet(_fromUtf8("background-color: rgb(255, 0, 0);"))
    flag.setAlignment(QtCore.Qt.AlignCenter)
    flag.setText(_translate("Dialog", text, None))
    return flag

def create_button(Dialog, x, y, text):
    qbutton = QtGui.QPushButton(Dialog)
    qbutton.setGeometry(QtCore.QRect(x, y, 40, 16))
    qbutton.setText(_translate("Dialog", text, None))
    return qbutton

def create_record(i,diz):
    record = {}

    record['frame'] = QtGui.QFrame()
    record['frame'].setFrameShape(1)
    record['frame'].setFixedSize(960,20)
    create_label(record['frame'], 8, 2, 15, 18, str(i))
    record['Base']=int(diz['Base'])
    record['TPM']=diz['TPM']
    record['RX']=diz['RX']
    create_label(record['frame'], 35, 2, 30, 18, str(int(diz['Base'])))
    if not diz['HC']=="":
        create_label(record['frame'], 82, 2, 30, 18, str(int(diz['HC'])))
    if not diz['ROX']=="":
        create_label(record['frame'], 120, 2, 30, 18, str(int(diz['ROX'])))
    if not diz['Ribbon']=="":
        create_label(record['frame'], 155, 2, 30, 18, str(int(diz['Ribbon'])))
        create_label(record['frame'], 190, 2, 30, 18, str(int(diz['Fibre'])))
    create_label(record['frame'], 230, 2, 50, 18, str(diz['Colour']))
    if not diz['TPM']=="":
        create_label(record['frame'], 290, 2, 20, 18, str(int(diz['TPM'])))
        create_label(record['frame'], 320, 2, 20, 18, str(int(diz['RX'])))
    create_flag(record['frame'], 380, 2, "yellow", "N")
    create_flag(record['frame'], 410, 2, "green", "S")
    create_flag(record['frame'], 440, 2, "green", "D")
    create_flag(record['frame'], 470, 2, "green", "F")
    create_flag(record['frame'], 500, 2, "green", "G")
    create_flag(record['frame'], 530, 2, "green", "C")
    create_flag(record['frame'], 560, 2, "green", "P")
    create_flag(record['frame'], 590, 2, "green", "R")
    record['add']=create_button(record['frame'], 650, 2, "Plot")
    record['att']=create_button(record['frame'], 690, 2, "Att")

    return record


def create_plot_record(ant, cells):
    record = {}
    diz = [x for x in cells if x['Base']==ant][0]
    record['frame'] = QtGui.QFrame()
    record['frame'].setFrameShape(1)
    record['frame'].setFixedSize(960,20)
    #create_label(record['frame'], 8, 2, 15, 18, str(i))
    record['Base']=int(diz['Base'])
    record['TPM']=diz['TPM']
    record['RX']=diz['RX']
    record['HC']=diz['HC']
    create_label(record['frame'], 35, 2, 30, 18, str(int(diz['Base'])))
    if not diz['HC']=="":
        create_label(record['frame'], 82, 2, 30, 18, str(int(diz['HC'])))
    if not diz['ROX']=="":
        create_label(record['frame'], 120, 2, 30, 18, str(int(diz['ROX'])))
    if not diz['Ribbon']=="":
        create_label(record['frame'], 155, 2, 30, 18, str(int(diz['Ribbon'])))
        create_label(record['frame'], 190, 2, 30, 18, str(int(diz['Fibre'])))
    create_label(record['frame'], 230, 2, 50, 18, str(diz['Colour']))
    if not diz['TPM']=="":
        create_label(record['frame'], 290, 2, 20, 18, str(int(diz['TPM'])))
        create_label(record['frame'], 320, 2, 20, 18, str(int(diz['RX'])))
    create_flag(record['frame'], 380, 2, "yellow", "N")
    create_flag(record['frame'], 410, 2, "green", "S")
    create_flag(record['frame'], 440, 2, "green", "D")
    create_flag(record['frame'], 470, 2, "green", "F")
    create_flag(record['frame'], 500, 2, "green", "G")
    create_flag(record['frame'], 530, 2, "green", "C")
    create_flag(record['frame'], 560, 2, "green", "P")
    create_flag(record['frame'], 590, 2, "green", "R")
    record['Color']=create_color_list(record['frame'],640,1)

    return record

def create_color_list(frame,x,y):
    listacolori=QtGui.QComboBox(frame)
    listacolori.addItem("blue")
    listacolori.addItem("green")
    listacolori.addItem("red")
    listacolori.addItem("black")
    listacolori.addItem("yellow")
    listacolori.addItem("cyan")
    listacolori.setGeometry(QtCore.QRect(x, y, 100, 18))
    listacolori.setFont(QtGui.QFont("Ubuntu", 9))
    return listacolori



def font_bold():
    font = QtGui.QFont()
    font.setBold(True)
    font.setWeight(75)
    return font
def font_normal():
    font = QtGui.QFont()
    return font



class AAVS(QtGui.QMainWindow):

    """ Main UI Window class """

    # Signal for Slots
    plot_signal = QtCore.pyqtSignal()
    #jig_pm_signal = QtCore.pyqtSignal()
    #antenna_test_signal = QtCore.pyqtSignal()

    def __init__(self, uiFile, flag_debug=False, opt_tpms=""):

        """ Initialise main window """
        super(AAVS, self).__init__()

        # Load window file
        self.mainWidget = uic.loadUi(uiFile)
        self.setCentralWidget(self.mainWidget)
        self.setWindowTitle("AAVS")
        self.resize(1100,680)

        # Flag for thread
        self.process_live_enabled = False

        self.debug = flag_debug
        print "The program is running with flag DEBUG set to:",self.debug 

        if not opt_tpms == "":
            self.tpm_list = opt_tpms.split(",")
        else:
            self.tpm_list = []
        #self.pic_ska = QtGui.QLabel(self.mainWidget.qtab_conf)
        #self.pic_ska.setGeometry(590, 300, 480, 200)
        #self.pic_ska.setPixmap(QtGui.QPixmap(os.getcwd() + "/pic/ska_inaf_logo2.jpg"))
        #self.mainWidget.qframe_ant_rms.hide()

        self.cells = []
        self.TPMs=[]

        self.gb = self.mainWidget.cb_group.isChecked()
        self.gb_not = self.mainWidget.cb_not.isChecked()
        self.gb_single = self.mainWidget.cb_single.isChecked()
        self.gb_double = self.mainWidget.cb_double.isChecked()
        self.gb_ferrite = self.mainWidget.cb_ferrite.isChecked()
        self.gb_gnd = self.mainWidget.cb_gnd.isChecked()
        self.gb_base = self.mainWidget.cb_base.isChecked()
        self.gb_debug = self.mainWidget.cb_debug.isChecked()
        self.switchtox = False
        self.switchtoy = False
        self.load_events()
        self.loadAAVSdata()
        self.countgroups()
        self.fft_nsamples = 1024
        self.log = True

        if self.log:
            for tpm in self.TPMs:
                directory = LOG_PATH + "/" + tpm['IP']
                if not os.path.exists(directory):
                    os.makedirs(directory)

        self.mapPlot = MapPlot(self.mainWidget.plotWidgetMap)
        self.mapPlot.plotClear()
        self.runMap()
        self.mapPlot.canvas.mpl_connect('button_press_event', self.onclick)
        self.mapPlot.canvas.mpl_connect('motion_notify_event', self.onmotion)
        self.show()

        self.initPlotList()
        self.updatePlotList()

        self.spectraPlot = MiniPlots(self.mainWidget.plotWidgetSpectra, 16, dpi=92)
        #self.spectraPlot.plotClear()

        self.process_livePlot = Thread(target=self.read_tpm_data)

    def reloadAAVS(self):
        self.loadAAVSdata()
        self.countgroups()

    def runMap(self,reload_val=True):
        self.mapPlot.plotClear()
        print "RUN MAP"
        self.map_polx = self.mainWidget.cb_polx.isChecked()
        self.map_poly = self.mainWidget.cb_poly.isChecked()
        if self.map_poly:
            self.pol=1
        else:
            self.pol=0

        if self.mainWidget.cb_rms.isChecked():
            self.map_meas = 'RMS'
        else:
            self.map_meas = 'DBM'
        self.map_dBm = self.mainWidget.cb_dBm.isChecked()
        self.map_base = self.mainWidget.cb_base.isChecked()
        if self.mainWidget.cb_tpm.isChecked():
            for i in map_tpm_range:
                self.color_ant_of_tpm(i[0],i[2])
        else:
            self.mapPlot.plotMap(self.cells, marker='8', markersize=12, color='k')

        if self.mainWidget.cb_group.isChecked():
            if self.mainWidget.cb_double.isChecked():
                dbl=[a for a in self.cells if not a['D_C']==""]
                self.mapPlot.plotMap(dbl, marker='8', markersize=15, color='k')
                print("Count Double_Clean: %d"%(len(dbl)))

            if self.mainWidget.cb_single.isChecked():
                dbl=[a for a in self.cells if not a['S_C']==""]
                self.mapPlot.plotMap(dbl, marker='8', markersize=15, color='k')
                print("Count Single_Clean: %d"%(len(dbl)))

            if self.mainWidget.cb_not.isChecked():
                dbl=[a for a in self.cells if not a['N_C']==""]
                self.mapPlot.plotMap(dbl, marker='8', markersize=15, color='k')
                print("Count Not_Clean: %d"%(len(dbl)))

            if self.mainWidget.cb_ferrite.isChecked():
                dbl=[a for a in self.cells if not a['Fer']==""]
                self.mapPlot.plotMap(dbl, marker='8', markersize=15, color='k')
                print("Count Ferrite: %d"%(len(dbl)))

            if self.mainWidget.cb_debug.isChecked():
                dbl=[a for a in self.cells if not a['Debug']==""]
                self.mapPlot.plotMap(dbl, marker='8', markersize=15, color='k')
                print("Count Debug: %d"%(len(dbl)))

            if self.mainWidget.cb_gnd.isChecked():
                dbl=[a for a in self.cells if not a['HC_GND']==""]
                self.mapPlot.plotMap(dbl, marker='8', markersize=15, color='k')
                print("Count GND Jacket: %d"%(len(dbl)))

            if self.mainWidget.cb_off.isChecked():
                dbl = [a for a in self.cells if not a['DC'] == ""]
                self.mapPlot.plotMap(dbl, marker='8', markersize=15, color='k')
                print("Count OFF: %d" % (len(dbl)))

            if self.mainWidget.cb_base.isChecked():
                self.mapPlot.plotMap(self.cells, marker='8', markersize=15, color='k')
                print("Count Basements: %d" % (len(dbl)))

            if self.mainWidget.cb_ant.isChecked():
                dbl = [a for a in self.cells if not a['HC'] == ""]
                self.mapPlot.plotMap(dbl, marker='8', markersize=15, color='k')
                print("Count Antennas: %d" % (len(dbl)))

            if self.mainWidget.cb_gold.isChecked():
                dbl = [a for a in self.cells if not a['Gold'] == ""]
                self.mapPlot.plotMap(dbl, marker='8', markersize=15, color='k')
                print("Count Gold: %d" % (len(dbl)))

        #self.mapPlot.plotMap(self.cells, marker='8', markersize=10, color='w')

        spente=[a for a in self.cells if a['DC']=="OFF"]
        #self.mapPlot.plotMap(spente, marker='+', markersize=11, color='k')
        if self.mainWidget.cb_basename.isChecked():
            self.mapPlot.printBase(self.cells)
        else:
            deployed = [a for a in self.cells if not a['ROX'] == ""]
            self.mapPlot.plotMap(deployed, marker='+', markersize=11, color='k')


        if self.mainWidget.cb_rms.isChecked() or self.mainWidget.cb_dBm.isChecked():
            for tpm in self.TPMs:
                if len(tpm['ANTENNE'])>0:
                    if reload_val:
                        self.freqs, self.spettro, rawdata, rms, dbm = get_raw_meas(tpm, nsamples=self.fft_nsamples, debug=self.debug)
                        tpm['RMS'] = rms
                        tpm['DBM'] = dbm
                #print len(tpm['ANTENNE']), len(rms)
                for j in range(len(tpm['ANTENNE'])):
                    if (tpm['ANTENNE'][j]['DC']==""):
                        x = float(str(tpm['ANTENNE'][j]['East']).replace(",", "."))
                        y = float(str(tpm['ANTENNE'][j]['North']).replace(",", "."))
                        tpm['ANTENNE'][j]['RMS-X']=tpm['RMS'][(j*2)]
                        tpm['ANTENNE'][j]['RMS-Y']=tpm['RMS'][(j*2)+1]
                        tpm['ANTENNE'][j]['DBM-X']=tpm['DBM'][(j*2)]
                        tpm['ANTENNE'][j]['DBM-Y']=tpm['DBM'][(j*2)+1]
                        if (tpm['RMS'][(j*2)] > 105) or (tpm['RMS'][(j*2)+1] > 105):
                            self.plotOscilla( x, y)
                        else:
                            if self.map_meas=="RMS":
                                self.mapPlot.oPlot(x, y, marker='8', markersize=10, color=rms_color(tpm['RMS'][(j*2)+self.pol]))
                            else:
                                self.mapPlot.oPlot(x, y, marker='8', markersize=10, color=rms_color(tpm['DBM'][(j*2)+self.pol]))
                        if self.log:
                            self.logMeas(tpm['IP'], tpm['ANTENNE'][j])
    #                    print("Plotting Antenna "+str(tpm['ANTENNE'][j]['Base'])+" (RX: "+str(tpm['ANTENNE'][j]['RX'])+", J:"+str(j)+", Colour: "+tpm['ANTENNE'][j]['Colour']+") with RMS %4.1f"%(rms[(j*2)+self.pol])+" and power of %4.1f dBm"%(dbm[(j*2)+self.pol]))
                        print("Plotting Antenna "+str(tpm['ANTENNE'][j]['Base'])+" with RMS %4.1f"%(tpm['RMS'][(j*2)+self.pol])+" and power of %4.1f dBm"%(tpm['DBM'][(j*2)+self.pol]))
                    else:
                        print("Plotting Antenna "+str(tpm['ANTENNE'][j]['Base'])+" Powered OFF")
                        x = float(str(tpm['ANTENNE'][j]['East']).replace(",", "."))
                        y = float(str(tpm['ANTENNE'][j]['North']).replace(",", "."))
                        self.mapPlot.oPlot(x, y, marker='8', markersize=10, color='k')

        self.mapPlot.plotMap(spente, marker='8', markersize=8, color='k')
        noant = [a for a in self.cells if a['HC'] == ""]
        self.mapPlot.plotMap(noant, marker='8', markersize=12, color='w')
        self.annot = self.mapPlot.canvas.ax.annotate("", xy=(0, 0), xytext=(20, 20), textcoords="offset points",
                            bbox=dict(boxstyle="round", fc="w"),
                            arrowprops=dict(arrowstyle="->"))
        self.annot.set_visible(False)

    def plotOscilla(self, x, y):
        self.mapPlot.oPlot(x, y, marker='8', markersize=14, color='k')
        self.mapPlot.oPlot(x, y, marker='8', markersize=12, color='y')
        self.mapPlot.oPlot(x, y, marker='8', markersize=9, color='k')
        self.mapPlot.oPlot(x, y, marker='8', markersize=6, color='y')
        self.mapPlot.oPlot(x, y, marker='8', markersize=3, color='k')
        self.mapPlot.oPlot(x, y, marker='8', markersize=2, color='y')


    def logMeas(self, ip, ant):
        filename = LOG_PATH+"/"+ip+"/ANTENNA-%03d.tsv"%(int(ant['Base']))
        if os.path.exists(filename):
            append_write = 'a' # append if already exists
        else:
            append_write = 'w' # make a new file if not
        #print filename
        with open(filename, append_write) as f:
            f.write(datetime.datetime.strftime(datetime.datetime.utcfromtimestamp(time.time()),"%Y-%m-%d %H:%M:%S\t")+"%4.1f\t%4.1f\t%4.1f\t%4.1f\n"%(ant['RMS-X'],ant['RMS-Y'],ant['DBM-X'],ant['DBM-Y']))

    def saveMap(self):
        print "SAVE MAP"

    def plotMap(self):
        print "PLOT MAP"

    def loadAAVSdata(self):
        if not self.debug:
            try:
                self.cells = read_from_google()
                print "\nSuccessfully connected to the online google spreadsheet!\n\n"

            except httplib2.ServerNotFoundError:
                print("\nUnable to find the server at accounts.google.com.\n\nContinuing with local file: %s\n"%(EX_FILE))
                self.cells = read_from_local()
                print "done!"
            tpms = ip_scan()
        else:
            if os.path.isfile(EX_FILE):
                print("\nReading local file: %s\n" % (EX_FILE))
                self.cells = read_from_local(EX_FILE)
            else:
                print("\nReading local file: %s\n" % (EX_FILE_AAVS))
                self.cells = read_from_local(EX_FILE_AAVS)
            tpms=["10.0.10."+str(a) for a in xrange(1,17)]

        if not self.tpm_list == []:
            tpms =  self.tpm_list
        self.TPMs=[]
        for i in tpms:
            tpm = {}
            tpm['TPM'] = TPM(ip=i, port=10000, timeout=1)
            tpm['IP']  = i
            tpm['ANTENNE'] = [x for x in self.cells if x['TPM']==int(i.split(".")[-1])]
            self.TPMs += [tpm]
        print("Loaded objects of %d TPM"%(len(self.TPMs)))

        mygroupbox = QtGui.QGroupBox()
        myform = QtGui.QFormLayout()
        myform.setVerticalSpacing(0)
        myform.setHorizontalSpacing(0)
        self.records = []
        for i in xrange(len(self.cells)):
            self.records += [create_record(i, self.cells[i])]
            ant=int(self.records[i]['Base'])
            self.records[i]['add'].clicked.connect(lambda status, g=ant: self.addPlot(g))
            self.records[i]['att'].clicked.connect(lambda status, g=ant: self.openPreadu(g))

            myform.addRow(self.records[i]['frame'])
            #myform.addRow(QtGui.QLabel("Ciao"))
        mygroupbox.setLayout(myform)
        self.mainWidget.qscroll_aavs.setWidget(mygroupbox)
        self.mainWidget.qscroll_aavs.setWidgetResizable(True)
        self.mainWidget.qscroll_aavs.setFixedHeight(221)
        #print("len(self.cells)=%d"%len(self.cells))

    def initPlotList(self):
        self.plotRecords = QtGui.QGroupBox()
        myform2 = QtGui.QFormLayout()
        myform2.setVerticalSpacing(0)
        myform2.setHorizontalSpacing(0)
        #myform.addRow()
        self.plotRecords.setLayout(myform2)
        self.mainWidget.qscroll_plot.setWidget(self.plotRecords)
        self.mainWidget.qscroll_plot.setWidgetResizable(True)
        self.mainWidget.qscroll_plot.setFixedHeight(120)
        self.plotList = []

    def addPlot(self, ant, color=0):
        print "Adding to the plot list Antenna # %d"%(ant)

        rec = create_plot_record(ant, self.cells)
        rec['Color'].setCurrentIndex(color)
        rec['Color'].currentIndexChanged.connect(lambda state, g=rec: self.change_color(g))
        myform=self.plotRecords.layout()

        myform.addRow(rec['frame'])
        self.plotRecords.setLayout(myform)
        self.plotList += [[rec['HC'],rec['TPM'],rec['RX'],color]]
        pass

    def updatePlotList(self):
        self.mainWidget.qcombo_plotList.clear()
        lista=sorted(os.listdir(PATH_PLOT_LIST))
        for i in lista:
            self.mainWidget.qcombo_plotList.addItem(i.split(".")[0])
        pass


    def change_color(self,r):
        for i in xrange(len(self.plotList)):
            if self.plotList[i][0] == r['HC']:
                self.plotList[i] = [r['HC'],r['TPM'],r['RX'],r['Color'].currentIndex()]
                break


    def openPreadu(self, record):
        pass


    def clearPlotList(self):
        self.initPlotList()
        self.plotList = []
        print "Plot List deleted!"

    def savePlotList(self):
        if not os.path.exists(PATH_PLOT_LIST):
            os.makedirs(PATH_PLOT_LIST)
        okscrivi = True
        if os.path.isfile(PATH_PLOT_LIST+self.mainWidget.qtext_listname.text()+".list"):
            result = QtGui.QMessageBox.question(self,
                          "Confirm Overwrite...",
                          "Are you sure you want to overwrite the existing file \""+self.mainWidget.qtext_listname.text()+"\" ?",
                          QtGui.QMessageBox.Yes| QtGui.QMessageBox.No)

            if result == QtGui.QMessageBox.Yes:
                okscrivi = True
            else:
                okscrivi = False
        if okscrivi:
            with open(PATH_PLOT_LIST+self.mainWidget.qtext_listname.text()+".list","w") as f:
                print self.plotList
                for (ant,tpm,rx,color) in self.plotList:
                    print ant,tpm,rx,color
                    f.write(str(int(ant))+","+str(int(tpm))+","+str(int(rx))+","+str(int(color))+"\n")
            print "Plot List \""+self.mainWidget.qtext_listname.text()+"\" saved!"
        self.updatePlotList()

    def loadPlotList(self):
        nomelista = easygui.fileopenbox(msg='Please select the Plot List', default=".plotlists/*.list")
        try:
            with open(nomelista, "r") as f:
                a=f.readlines()
        except:
            print "Not able to load the list file!"
            pass
        self.clearPlotList()
        for i in a:
            print i
            self.addPlot(int(i.strip().split(",")[0]), int(i.strip().split(",")[3]))

        print "Plot List loaded."

    def change_poly(self):
        if not self.switchtox:
            print self.mainWidget.cb_polx.isChecked(), self.mainWidget.cb_poly.isChecked(),
            self.switchtoy = True
            if self.mainWidget.cb_poly.isChecked():
                self.poly = True
                self.polx = False
                self.mainWidget.cb_polx.setChecked(False)
            elif not self.mainWidget.cb_polx.isChecked():
                self.poly = True
                self.mainWidget.cb_poly.setChecked(True)
            self.switchtoy = False
            print self.mainWidget.cb_polx.isChecked(),self.mainWidget.cb_poly.isChecked()

    def change_polx(self):
        if not self.switchtoy:
            print self.mainWidget.cb_polx.isChecked(), self.mainWidget.cb_poly.isChecked(),
            self.switchtox = True
            if self.mainWidget.cb_polx.isChecked():
                self.polx = True
                self.poly = False
                self.mainWidget.cb_poly.setChecked(False)
            elif not self.mainWidget.cb_poly.isChecked():
                 self.polx = True
                 self.mainWidget.cb_polx.setChecked(True)
            self.switchtox=False
            print self.mainWidget.cb_polx.isChecked(),self.mainWidget.cb_poly.isChecked()


    def onclick(self, event):
        #print("APRITI!!!")
        if event.dblclick and not event.xdata == None:
            self.popwd = QtGui.QDialog()
            self.popw = AAVS_SNAP_Dialog()
            self.popw.setupUi(self.popwd)
            self.popwd.show()

            self.popwPlot = MiniPlots(self.popw.frame, 1, dpi=85)
            #print event.xdata, event.ydata
            if event.button == 1:
                #print len(self.cells)
                sel = [x for x in self.cells if ((x['East'] > event.xdata - 0.6) and (x['East'] < event.xdata + 0.6))]
                res = [x for x in sel if ((x['North'] > event.ydata - 0.6) and (x['North'] < event.ydata + 0.6))]
                if len(res) == 1:
                    board = {}
                    print "Selected antenna", int(res[0]['Base'])  # , len(TPMs)#,res[0]['East'], res[0]['North']
                    for i in range(len(self.TPMs)):
                        # for x in TPMs[i]['ANTENNE']:
                        #    print x['Base'], " ",
                        # print
                        if len([x for x in self.TPMs[i]['ANTENNE'] if int(x['Base']) == int(res[0]['Base'])]) > 0:
                            board['IP'] = self.TPMs[i]['IP']
                            board['TPM'] = self.TPMs[i]['TPM']
                            board['ANTENNE'] = [x for x in self.TPMs[i]['ANTENNE'] if int(x['Base']) == int(res[0]['Base'])]
                            # print board['ANTENNE'], board['IP']
                    if not board == {}:
                        if len(board['ANTENNE']) == 1:
                            self.freqs, self.spettro , rawdata , rms, rfpower = get_raw_meas(board, nsamples=1024, debug=self.debug)
                            print "Returned", len(self.spettro), len(self.spettro[0])
                            self.popwPlot.plotCurve(self.freqs, self.spettro[(int(board['ANTENNE'][0]['RX']) - 1) * 2], 0,colore='b', label="Pol X",yAxisRange=[-80,-20])
                            self.popwPlot.plotCurve(self.freqs, self.spettro[(int(board['ANTENNE'][0]['RX']) - 1) * 2 + 1], 0, colore='g', label="Pol Y",yAxisRange=[-80,-20])
                            self.popwPlot.updatePlot()
                            self.popw.qlabel_antnum.setText("Antenna Base # " + str(int(board['ANTENNE'][0]['Base'])))
                            self.popw.qlabel_hc.setText("Hybrid Cable: " + str(int(board['ANTENNE'][0]['HC'])))
                            self.popw.qlabel_rox.setText("Roxtec: " + str(int(board['ANTENNE'][0]['ROX'])))
                            self.popw.qlabel_rib.setText("Ribbon: " + str(int(board['ANTENNE'][0]['Ribbon'])))
                            self.popw.qlabel_fib.setText("Fibre: " + str(int(board['ANTENNE'][0]['Fibre'])))
                            self.popw.qlabel_col.setText("Colour: " + str(board['ANTENNE'][0]['Colour']))
                            self.popw.qlabel_tpm.setText("TPM: " + str(int(board['ANTENNE'][0]['TPM'])))
                            self.popw.qlabel_rx.setText("RX: " + str(int(board['ANTENNE'][0]['RX'])))

                elif len(res) > 2:
                    print "Search provides more than one result (found %d candidates)" % (len(res))
                else:
                    # print "Double clicked on x:%4.2f and y:%4.2f, no antenna found here!"%(event.xdata,event.ydata)
                    pass
            else:
                pass
        else:
            if (not event.xdata == None) and (not event.ydata == None):
                #print event.xdata, event.ydata
                for tpms in map_tpm_range:
                    if event.xdata > tpms[1][0] and event.xdata < tpms[1][2]:
                        if event.ydata > tpms[1][3] and event.ydata < tpms[1][1]:
                            print "Clicked on ",tpms[0]
                            self.color_ant_of_tpm(tpms[0],tpms[2])
                            deployed = [a for a in self.cells if not a['ROX'] == "" and a['TPM']==int(tpms[0][4:])]
                            self.mapPlot.plotMap(deployed, marker='+', markersize=11, color='k')
                            break

    def update_annot(self, x,y,text):

        pos = (x,y)
        self.annot.xy = pos
        self.annot.set_text(text)
        self.annot.get_bbox_patch().set_facecolor('w')
        #self.annot.get_bbox_patch().set_alpha(0.4)

    def mouseonbase(self, x, y):
        if (x - 0.5 < 0) and (x + 0.5 > 0) and (y - 0.5 < 0) and (y + 0.5 >
                                                                      0):
            return "APIU"
        else:
            for a in self.cells:
                if ((a['East'] > x - 0.6) and (a['East'] < x + 0.6)):
                    if ((a['North'] > y - 0.6) and (a['North'] < y + 0.6)):
                        if not a['HC']=="":
                            return "Base: "+str(a['Base'])+"\nHC: "+str(int(a['HC']))+", Rox: "+str(int(a['ROX']))+"\nRib: "+str(int(a['Ribbon']))+", F: "+str(int(a['Fibre']))
                        else:
                            return "Base: "+str(a['Base'])+"\nNo Antenna\ndeployed"
            return ""

    def onmotion(self, event):
        #pass
        #print event.xdata, event.ydata
        vis = self.annot.get_visible()
        if event.inaxes == self.mapPlot.canvas.ax:
            cont, ind = self.mapPlot.canvas.ax.contains(event)
            hc = self.mouseonbase(event.xdata, event.ydata)
            if ((cont) and (not hc=="")):
                self.update_annot(event.xdata, event.ydata,hc)
                self.annot.set_visible(True)
                self.mapPlot.updatePlot()
                #self.mapPlot.canvas.ax.draw_idle()
            else:
                if vis:
                    self.annot.set_visible(False)
                    self.mapPlot.updatePlot()
                    #self.mapPlot.canvas.draw_idle()

        #print("APRITI!!!")

    def color_ant_of_tpm(self, tpm, colore):
        t=int(tpm[4:])
        #print t
        #for t in self.TPMs:
        antenne = [a for a in self.cells if a['TPM']==t]
        #print antenne, self.cells
        if not antenne==[]:
            self.mapPlot.plotMap(antenne, marker='8', markersize=15, color=colore)

    def plotSpectra(self):
        self.lista_spettri = []
        with open(PATH_PLOT_LIST+self.mainWidget.qcombo_plotList.currentText()+".list", "r") as f:
            a = f.readlines()
        for i in a:
            self.lista_spettri += [i.strip().split(",")[0], int(i.strip().split(",")[1])]
        #print self.lista_spettri
        print "Plotta :D"

    def saveSpectra(self):
        print "Save :D"


    def load_events(self):
        self.mainWidget.button_reload.clicked.connect(lambda: self.reloadAAVS())
        self.mainWidget.button_runMap.clicked.connect(lambda: self.runMap())
        self.mainWidget.button_saveMap.clicked.connect(lambda: self.saveMap())
        self.mainWidget.qbutton_DeleteList.clicked.connect(lambda: self.clearPlotList())
        self.mainWidget.qbutton_SaveList.clicked.connect(lambda: self.savePlotList())
        self.mainWidget.qbutton_LoadList.clicked.connect(lambda: self.loadPlotList())
        self.mainWidget.cb_poly.stateChanged.connect(lambda: self.change_poly())
        self.mainWidget.cb_polx.stateChanged.connect(lambda: self.change_polx())
        #self.mainWidget.qbutton_plotStart.clicked.connect(lambda: self.plotSpectra())
        self.mainWidget.qbutton_plotSave.clicked.connect(lambda: self.saveSpectra())
        self.mainWidget.cb_group.stateChanged.connect(lambda: self.select_group_en())
        self.mainWidget.cb_double.stateChanged.connect(lambda: self.runMap(False))
        self.mainWidget.cb_single.stateChanged.connect(lambda: self.runMap(False))
        self.mainWidget.cb_not.stateChanged.connect(lambda: self.runMap(False))
        self.mainWidget.cb_ferrite.stateChanged.connect(lambda: self.runMap(False))
        self.mainWidget.cb_gnd.stateChanged.connect(lambda: self.runMap(False))
        self.mainWidget.cb_debug.stateChanged.connect(lambda: self.runMap(False))

        self.mainWidget.qbutton_plotStart.clicked.connect(lambda: self.ant_enable())
        self.mainWidget.qbutton_plotSave.clicked.connect(lambda: self.savePlot())
        self.mainWidget.qcombo_plotList.currentIndexChanged.connect(self.remakeList)
        self.mainWidget.rb_single.toggled.connect(self.pltsingle)
        self.mainWidget.rb_multi.toggled.connect(self.pltmulti)

    def savePlot(self):
        print "Save Plot: To be implemented"
        pass

    def pltsingle(self):
        self.mainWidget.rb_polx.setEnabled(True)
        self.mainWidget.rb_poly.setEnabled(True)
        pass

    def pltmulti(self):
        self.mainWidget.rb_polx.setEnabled(False)
        self.mainWidget.rb_poly.setEnabled(False)
        pass


    def select_group_en(self):
        if self.mainWidget.cb_group.isChecked():
            self.mainWidget.cb_double.setEnabled(True)
            self.mainWidget.cb_single.setEnabled(True)
            self.mainWidget.cb_not.setEnabled(True)
            self.mainWidget.cb_ferrite.setEnabled(True)
            self.mainWidget.cb_gnd.setEnabled(True)
            self.mainWidget.cb_debug.setEnabled(True)
            self.mainWidget.cb_base.setEnabled(True)
            self.mainWidget.cb_off.setEnabled(True)
            self.mainWidget.count_debug.setEnabled(True)
            self.mainWidget.count_sc.setEnabled(True)
            self.mainWidget.count_dbl.setEnabled(True)
            self.mainWidget.count_nc.setEnabled(True)
            self.mainWidget.count_fer.setEnabled(True)
            self.mainWidget.count_off.setEnabled(True)
            self.mainWidget.count_base.setEnabled(True)
            self.mainWidget.count_gnd.setEnabled(True)
            self.mainWidget.count_gold.setEnabled(True)
            self.mainWidget.count_ant.setEnabled(True)
        else:
            self.mainWidget.cb_double.setEnabled(False)
            self.mainWidget.cb_single.setEnabled(False)
            self.mainWidget.cb_not.setEnabled(False)
            self.mainWidget.cb_ferrite.setEnabled(False)
            self.mainWidget.cb_gnd.setEnabled(False)
            self.mainWidget.cb_debug.setEnabled(False)
            self.mainWidget.cb_base.setEnabled(False)
            self.mainWidget.cb_off.setEnabled(False)
            self.mainWidget.count_debug.setEnabled(False)
            self.mainWidget.count_sc.setEnabled(False)
            self.mainWidget.count_dbl.setEnabled(False)
            self.mainWidget.count_nc.setEnabled(False)
            self.mainWidget.count_fer.setEnabled(False)
            self.mainWidget.count_off.setEnabled(False)
            self.mainWidget.count_base.setEnabled(False)
            self.mainWidget.count_gnd.setEnabled(False)
            self.mainWidget.count_gold.setEnabled(False)
            self.mainWidget.count_ant.setEnabled(False)

    def select_dblcleaned(self):
        self.runMap(False)

    def countgroups(self):
        try:
            #print self.cells[0],self.cells[-1]
            dbl=[a for a in self.cells if not a['D_C']==""]
            self.mainWidget.count_dbl.setText("Count: "+str(len(dbl)))
            dbl=[a for a in self.cells if not a['S_C']==""]
            self.mainWidget.count_sc.setText("Count: "+str(len(dbl)))
            dbl=[a for a in self.cells if not a['N_C']==""]
            self.mainWidget.count_nc.setText("Count: "+str(len(dbl)))
            dbl=[a for a in self.cells if not a['Fer']==""]
            self.mainWidget.count_fer.setText("Count: "+str(len(dbl)))
            dbl=[a for a in self.cells if not a['Debug']==""]
            self.mainWidget.count_debug.setText("Count: "+str(len(dbl)))
            dbl=[a for a in self.cells if not a['HC_GND']==""]
            self.mainWidget.count_gnd.setText("Count: "+str(len(dbl)))
            dbl=[a for a in self.cells if not a['HC']==""]
            self.mainWidget.count_ant.setText("Count: "+str(len(dbl)))
            self.mainWidget.count_base.setText("Count: "+str(len(self.cells)))
            dbl=[a for a in self.cells if a['DC']=="OFF"]
            self.mainWidget.count_off.setText("Count: "+str(len(dbl)))
            dbl=[a for a in self.cells if not a['Gold']==""]
            self.mainWidget.count_gold.setText("Count: "+str(len(dbl)))
        except:
            print "ERROR while counting groups. Please check field names or the number of columns read in the spreadsheet."

    def remakeList(self):
        self.lista_spettri = []
        if not self.mainWidget.qcombo_plotList.currentText()=="":
            with open(PATH_PLOT_LIST+self.mainWidget.qcombo_plotList.currentText()+".list", "r") as f:
                a = f.readlines()
            for i in a:
                #self.lista_spettri += [[i.strip().split(",")[0], int(i.strip().split(",")[1])]]
                self.lista_spettri += [[int(float(a)) for a in i.strip().split(",")]]
            print "List: ",self.mainWidget.qcombo_plotList.currentText(),self.lista_spettri

    def ant_enable(self):
        if not self.process_livePlot.isAlive():
            self.process_livePlot.start()
        if not self.process_live_enabled:
            print "\nStart Antenna Measurements\n"
            self.process_live_enabled = True
            self.mainWidget.qbutton_plotStart.setText("Stop")
        else:
            print "\nStop Antenna Measurements\n"
            self.process_live_enabled = False
            self.mainWidget.qbutton_plotStart.setText("Start")

    def read_tpm_data(self):
        while True:
            if self.process_live_enabled:
                #print self.mainWidget.qcombo_plotList.currentText(), self.lista_spettri

                # Fare ciclo per le sole TPM nella lista
                tpml = np.unique(np.array([l[1] for l in self.lista_spettri]))
                #tpm = self.TPMs[0]
                self.data = []
                for tpm in self.TPMs:
                    if int(tpm['IP'].split(".")[-1]) in tpml:
                        print "Getting data of tile #",int(tpm['IP'].split(".")[-1]),"...",
                        self.freqs, self.spettro, rawdata, rms, dbm = get_raw_meas(tpm, nsamples=self.fft_nsamples,
                                                                           debug=self.debug)
                        for i in self.lista_spettri:
                            if i[1] == int(tpm['IP'].split(".")[-1]):
                                self.data += [{}]
                                self.data[-1]['hc'] = i[0]
                                self.data[-1]['tile'] = int(tpm['IP'].split(".")[-1])
                                self.data[-1]['data_x'] = self.spettro[i[2]*2]
                                self.data[-1]['data_y'] = self.spettro[i[2]*2+1]
                                self.data[-1]['freq'] = self.freqs
                                self.data[-1]['rx'] = i[2]
                                self.data[-1]['color'] = i[3]
                        print "done!"
                print "Emitting signal... (got data of ",str(len(tpml))," tiles {",tpml,"})"
                self.plot_signal.emit()
                cycle = 0.0
                while cycle<2:
                    time.sleep(0.5)
                    cycle = cycle + 0.5
            time.sleep(0.2)
            #if self.stopThreads:
            #    break

    def uplivePlot(self):
        #print len(self.spettro), len(self.spettro[0]),len(self.freqs)
        tempo = datetime.datetime.strftime(datetime.datetime.utcfromtimestamp(time.time()),"%Y-%m-%d   %H:%M:%S")
        self.mainWidget.qlabel_comment.setText("Last update on:  "+tempo)
        x_axis=[int(self.mainWidget.qtext_xmin.text()),int(self.mainWidget.qtext_xmax.text())]
        y_axis=[int(self.mainWidget.qtext_ymin.text()),int(self.mainWidget.qtext_ymax.text())]
        print "Updating Plot"
        if self.mainWidget.rb_multi.isChecked():
            self.spectraPlot = MiniPlots(self.mainWidget.plotWidgetSpectra, len(self.lista_spettri), dpi=92)
            for i in xrange(len(self.lista_spettri)):
                pltname  = "HC  "+str(self.data[i]['hc'])
                pltname += "    TPM " + str(self.data[i]['tile'])
                pltname += "    RX " + str(self.data[i]['rx'])
                self.spectraPlot.plotCurve(self.freqs[2:-2], self.data[i]['data_x'][2:-2], i, colore='b',title=pltname, xAxisRange=x_axis, yAxisRange=y_axis)
                self.spectraPlot.plotCurve(self.freqs[2:-2], self.data[i]['data_y'][2:-2], i, colore='g')
        else:
            self.spectraPlot = MiniPlots(self.mainWidget.plotWidgetSpectra, 1, dpi=92)
            labels = []
            #patches = []
            for i in xrange(len(self.lista_spettri)):
                labels  += ["HC  "+str(self.data[i]['hc'])]
                #patches += [mpatches.Patch(color=COLORS[self.data[i]['color']], label=labels[-1])]
                pltname = labels[-1] + "    TPM " + str(self.data[i]['tile'])
                pltname += "    RX " + str(self.data[i]['rx'])
                if self.mainWidget.rb_polx.isChecked():
                    self.spectraPlot.plotCurve(self.freqs[2:-2], self.data[i]['data_x'][2:-2], 0, colore=COLORS[self.data[i]['color']], xAxisRange=x_axis, yAxisRange=y_axis, label=labels[-1], title='Pol X', titlesize=13)
                else:
                    self.spectraPlot.plotCurve(self.freqs[2:-2], self.data[i]['data_y'][2:-2], 0, colore=COLORS[self.data[i]['color']], xAxisRange=x_axis, yAxisRange=y_axis, label=labels[-1], title='Pol Y', titlesize=13)
                print "Plot: "+labels[-1]
                #self.spectraPlot.plotCurve(self.freqs[2:-2], self.data[i]['data_y'][2:-2], i, colore='g')
            self.spectraPlot.canvas.ax[0].legend(labels,loc=2,prop={'size':9})
            #self.spectraPlot.canvas.ax[0].legend(handles=patches)
        self.spectraPlot.updatePlot()
        pass


if __name__ == "__main__":
    parser = OptionParser()
    parser.add_option("-d", "--debug", action='store_true',
                      dest="debug",
                      default=False,
                      help="If set the program runs in debug mode")

    parser.add_option("-t", "--tpms",
                      dest="tpms",
                      default="",
                      help="If given do not scan network but use the given list")

    (options, args) = parser.parse_args()

    os.system("python ../config/setup.py")
    app = QtGui.QApplication(sys.argv)
    window = AAVS("aavs.ui",options.debug,options.tpms)

    window.plot_signal.connect(window.uplivePlot)
    #window.jig_pm_signal.connect(window.updateJIGpm)
    #window.antenna_test_signal.connect(window.updateAntennaTest)

    sys.exit(app.exec_())
